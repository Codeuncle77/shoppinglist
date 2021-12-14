from openpyxl import load_workbook

book = load_workbook('test.xlsx')
print(book.sheetnames)